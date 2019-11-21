from openpyxl import load_workbook
import csv
from datetime import datetime

data_to_csv = []

excel_workbook = load_workbook('cotizaciones.xlsx', data_only=True)
excel_sheet = excel_workbook.active
for i, row in enumerate(excel_sheet.iter_rows(min_row=9, max_row=4437, min_col=4, max_col=9, values_only=True)):
    data_to_csv.append(row)

with open('log_cotizacion.csv', mode='w', newline='') as csv_data:
    csv_writer = csv.writer(csv_data, delimiter=',')
    for data in data_to_csv:
        date = datetime.date(data[0])
        comp1 = data[1]
        ven1 = data[2]
        comp2 = data[4]
        ven2 = data[5]

        csv_writer.writerow([date, comp1, ven1, comp2, ven2])
